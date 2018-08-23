import xlsxwriter
import openpyxl
import csv



def CalculateColour(Percent):
    if(Percent <= red_threshold):
        return "red"
    elif (Percent <= amber_threshold):
        return "amber"
    elif (Percent <= green_threshold):
        return "green"
    else:
        return "black"

class Learning_Objective:
    def __init__(self):
        self.LO_id = ""
        self.num_questions = 0
        self.score = 0
        self.percent = 0
        self.colour = ""
        
class Student:

    def __init__(self):
        self.student_id=0
        self.name='' #full name
        self.overall_score = 0
        self.Lo_answers = [] #a list of all the questions associated with this Learning objecive
  

class Learning_Objective_Question:

    def __init__(self):
        self.question_id="" #the id of the question
        self.answer=""      #the answer of the question if it  has text a answer

class Fronter2Doddle:

    def __init__(self):

        self.MODE_NORMAL = 1
        self.MODE_VERBOSE = 2
        self.MODE_PRODUCTION = 1
        
        self.VERSION = "2.0"
        self.error_string = ""

        self.error_flag = 0;
        self.red_threshold = 0
        self.amber_threshold = 0
        self.green_threshold = 0
        self.mode = self.MODE_NORMAL #MODE_VERBOSE   #no print statements in normal
        self.xlxs_workbook = None

        self.short_answer_list = []
        self.lo_questions = []
        self.question_ids = []
        self.num_of_questions = 0
        self.student_list = []


    def LogError(self, estring):
        self.error_string += estring + "\n"
        print(estring)
        
    def CalculateColour(self,Percent):
        if(Percent <= self.red_threshold):
            return "red"
        elif (Percent <= self.amber_threshold):
            return "amber"
        elif (Percent <= self.green_threshold):
            return "green"
        else:
            return "black"

    def log(self, log_str, imode = 0):
        if(imode == self.MODE_PRODUCTION):
            return
        elif(imode == self.MODE_NORMAL):
            print(log_str)
        elif(imode == self.MODE_VERBOSE):
            print(log_str)

    def CloseWorkbook(self):
        self.xlxs_workbook._archive.close()

    def LoadWorkbook(self, name_of_work_book):
        self.xlxs_workbook = openpyxl.load_workbook(name_of_work_book, use_iterators=True, read_only=True)

    def CalculateThreshholds(self,name_of_worksheet):
        thresh = self.xlxs_workbook.get_sheet_by_name(name_of_worksheet)
     
        #open the threshold file   
        for row in thresh.iter_rows():
            colour_name = row[0].value.strip().lower()
           
            if(str(colour_name) == "red"):
                self.red_threshold = int(row[1].value)
            elif(colour_name == "amber"):
                self.amber_threshold = int(row[1].value)
            elif(colour_name == "green"):
                self.green_threshold = int(row[1].value)
            else:
                print("Error: could not read threshold value")
                self.error_flag = 1
        self.log("Thresholds: Red:" + str(self.red_threshold) + " Amber:" + str(self.amber_threshold) + " Green:"+ str(self.green_threshold) + "\n")

    def GetQuestionIdsList(self,name_of_worksheet):
   
        #open the doddle questions csv file and store the ids in a list
        data = self.xlxs_workbook.get_sheet_by_name(name_of_worksheet)

        
        x = 0;
        for row in data.iter_rows():
            #skip the first row which has the coloumn headers
            if(x == 0):
                x = 1
                continue
            
            val = row[0].value.strip() #strip everything including the decimal place from the end of the learning objective string
            trim = len(val)-val.index(".")

            self.lo_question = val[0:len(val)-trim] #work out the learning objectibe i.e. D12.3 = D12
            self.lo_questions.append(self.lo_question)

            loq = Learning_Objective_Question() #store the id i.e. D12.3 along with the short text answer if applicable
            loq.question_id = str(val)

            text_answer = row[1].value #check for text answers

            loq.answer = "" #clear the answer
            
            if(text_answer != None):  #check if the cell is not null
                text_answer = text_answer.strip() #remove all whitespaces
                if(text_answer != ""): #check if we have text in the cell
                    loq.answer = text_answer.lower() #lowercase the string and store
                
            
            self.question_ids.append(loq) #add to the list of quesions 

        self.num_of_questions = len(self.question_ids)

                


    def CheckQuestionForTextAnswer(self,loq):
        for t  in self.question_ids:
            if(loq.question_id == t.question_id):
                if(loq.answer !=""):
                    return t
                else:
                    return None
        return None

    def CheckAnswer(self, correct, user):
        if ( set(correct.split()) == set(user.split()) ):
            return True
        else:
            return False

    def ProcessFronterTest(self, ifilename):

      
        f = open(ifilename,"r") #opens fronter test file
        
        id_index = 0
        name_index = 1
        total_score_index = 3
        start_of_answers_index = 4
                
        text_answers_list = []

    
        for line in f:
 
            #for each student
            line_elements = line.rstrip('\n').split('\t')

            #line elements    
            #print(line_elements)

            s = Student()
            s.student_id = line_elements[id_index]
            s.name = line_elements[name_index]
            s.name = s.name.split("-")[1]
            #s.name = s.name.replace('"', '')
            s.name = s.name.strip()

            current_lo = Learning_Objective()

            score_text_file = line_elements[total_score_index]

            total_score_calculated = 0
            lo_question = ""
            last_lo_question = ""
            current_lo_id = ""
            current_text_answer_index = 0
            #for all questions

            for x in range(0,self.num_of_questions):
                
                current_lo_id = self.lo_questions[x] #D1, D2, D4 etc
                current_loq = self.question_ids[x]   #D1.1, #D1.2 etc

                ta = self.CheckQuestionForTextAnswer(current_loq) #text answer from file

                if(ta != None): #this is a text answer
                    user_answer = line_elements[start_of_answers_index + x]
                    user_answer = user_answer.strip().lower()
                    correct_answer = ta.answer 
                    print ("tanswer id " + ta.question_id + "correct Answer is " + ta.answer + " user entered " + user_answer)
                    if (self.CheckAnswer(correct_answer, user_answer) == True):  
                        print("answer was right")
                        mark = 1
                    else:
                        print("answer was wrong")
                        mark = 0

                #note correct_answer = correct_answer.replace(" ", "") #WARNING removes all whitespaces from answer
                    
                else: #this is a regular answer with the 1/0 stored in the file
                    mark = float(line_elements[start_of_answers_index + x])
                    print("mark is " + str(mark))
                  
                    
                    
                total_score_calculated += mark

                #new learning objective
  
                if( last_lo_question != current_lo_id ):
                    last_lo_question = current_lo_id
                    if(x!=0):

                        current_lo.percent = (current_lo.score / current_lo.num_questions) * 100
                        current_lo.percent = int(current_lo.percent)
                        current_lo.colour = self.CalculateColour(current_lo.percent)
                        s.Lo_answers.append(current_lo)
                        current_lo = Learning_Objective()
                        current_lo.LO_id = self.lo_questions[x]
                    else:
                        current_lo.LO_id = self.lo_questions[x]

                    current_lo.num_questions += 1
                    current_lo.score += mark


                else:
                    current_lo.num_questions += 1
                    current_lo.score += mark

            #add the last learning objective
            current_lo.percent = (current_lo.score / current_lo.num_questions) * 100
            current_lo.percent = int(current_lo.percent)
            current_lo.colour = self.CalculateColour(current_lo.percent)

            s.Lo_answers.append(current_lo)
            
            #self.log("Calculating " + s.name,self.MODE_VERBOSE)
            self.student_list.append(s)



            #sanity check make sure the calculated score matched the score from the text file
            score_text_int = float((score_text_file) )
            s.overall_score = score_text_int
            if(total_score_calculated != score_text_int):
                self.LogError("ERROR 01 - the calculated score does not match score from the file")
                self.LogError("Student name " + s.name + " Student ID " + s.student_id)
                self.LogError(" Total score calculated " + str(total_score_calculated) + " Total score from file " + score_text_file )
                self.LogError(" Check that the number of questions in the test match the number of ids in the manifest file")
                self.error_flag = 1
                break;

        f.close()

        print()

        if(self.mode == self.MODE_VERBOSE):
            print("Output: ")
            print("--------------------------")
            for t in student_list:
                print("ID: " + t.student_id.replace('"', '') + " Name: " + t.name + " Score " + str(t.overall_score) )
                for k in t.Lo_answers:
                    print( k.LO_id + "- Questions " + str(k.num_questions) + " Marks " + str(k.score) + "  "  + str(k.percent) + "% -" + k.colour  )
                print()
                print("--------------------------")


    def WriteExcelOuputFile(self,ioutfilename):

        #write the xml file
        workbook = xlsxwriter.Workbook(ioutfilename)
        worksheet = workbook.add_worksheet()
        worksheet_breakdown = workbook.add_worksheet("Details")

        #cell background colours
        green_format = workbook.add_format()
        green_format.set_pattern(1)  
        green_format.set_bg_color('#99CC00')

        amber_format = workbook.add_format()
        amber_format.set_pattern(1)   
        amber_format.set_bg_color('#ff9900')

        red_format = workbook.add_format()
        red_format.set_pattern(1)   
        red_format.set_bg_color('#ff0000')

        black_format = workbook.add_format()
        black_format.set_pattern(1)   
        black_format.set_bg_color('#oo0000')


        worksheet_breakdown.write(0, 0, "Red",red_format)
        worksheet_breakdown.write(0, 1, self.red_threshold,red_format)

        worksheet_breakdown.write(0, 3, "Amber",amber_format)
        worksheet_breakdown.write(0, 4, self.amber_threshold,amber_format)

        worksheet_breakdown.write(0, 6, "Green",green_format)
        worksheet_breakdown.write(0, 7, self.green_threshold,green_format)

        row = 2
        col = 1

        for d in self.student_list[0].Lo_answers:
            worksheet.write(row, col, d.LO_id)
            worksheet_breakdown.write(row, col, d.LO_id)
            col = col+1

        row = 3
        col = 1

        worksheet.write(0, 0, "Ver " +  self.VERSION)

        worksheet.write(2, 0, "Name" )

        for t in self.student_list:
            worksheet.write(row, 0, t.name)
            worksheet_breakdown.write(row, 0, t.name)
            for l in t.Lo_answers:
                detail = str(l.score) + "/" + str(l.num_questions) + "-"  + str(l.percent) + "%"
                if(l.colour == "red"):
                    worksheet.write(row, col, l.colour, red_format)
                    worksheet_breakdown.write(row, col, detail,red_format)
                elif(l.colour == "amber"):
                    worksheet.write(row, col, l.colour, amber_format)
                    worksheet_breakdown.write(row, col, detail,amber_format)
                elif(l.colour == "green"):
                    worksheet.write(row, col, l.colour, green_format)
                    worksheet_breakdown.write(row, col, detail,green_format)
                else:
                    worksheet.write(row, col, l.colour, black_format)
                    worksheet_breakdown.write(row, col, detail,black_format)        
                
                col = col + 1
            col = 1
            row = row+1

        if(self.error_string != ""):
            worksheet_errors = workbook.add_worksheet("ERRORS")
            worksheet_errors.write(0, 0, "The following errors occured")
            worksheet_errors.write(1, 0, self.error_string ,red_format)
            
                 

        workbook.close()



    def ErrorCheck(self):
        if(self.error_flag == 0):
            print("...all is well")
            return False
        else:
            print("There were errors in the analysis")
            return True



if __name__ == "__main__":

    input_filename = "fronter_test.txt"

    fd = Fronter2Doddle()
    fd.log("Fronter2Doddle Ver " + fd.VERSION + "\n")

    fd.LoadWorkbook('Y7_COMP_AU2.xlsx')
    fd.CalculateThreshholds("Thresholds")
    fd.GetQuestionIdsList("Doddle Question Ids")
    fd.ProcessFronterTest(input_filename)
    fd.ErrorCheck()
    fd.WriteExcelOuputFile("class_file_new.xlsx")
    fd.CloseWorkbook()


    input("press any key to quit")




    

